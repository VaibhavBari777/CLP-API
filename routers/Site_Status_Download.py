from fastapi import APIRouter, HTTPException, BackgroundTasks
from fastapi.responses import FileResponse
from database import SessionLocal
from models import Data
from openpyxl import load_workbook
import shutil
import os
from uuid import uuid4
import math

router = APIRouter()

# Excel header → DB column mapping
EXCEL_COLUMN_MAP = {
    "Circle": "Circle_Name",
    "Project": "Project_Name",
    "Site ID": "Site_ID",
    "Vendor": "BTS_Vendor",
    "NSS ID": "NSS_ID",
    "MW OSM": "mw_osm",
    "Optics OSM": "optics_osm",
    "IP CR": "code_cr_details_ip_cr",
    "MW OSM Execution Status": "mw_osm_execution_status",
    "Optics OSM Execution Status": "optics_osm_execution_status",
    "IP CR Execution Status": "ip_cr_execution_status",
    "Tx E2E Readiness Status": "tx_e2e_readiness_status",
    "Site Status": "site_status",
    "Code E2E Status": "code_e2e_status"
}

SS_TEMPLATE_PATH = r"Provisioning\Sample_SITE STATUS PROVISIONING_Provisoning_Update_Template (1).xlsx"
TEMP_DIR = r"C:\CubeXo Internship\FastAPI\CLP_API\temp_downloads"


# 🔹 Cleanup temp file
def delete_file(path: str):
    if os.path.exists(path):
        os.remove(path)


# 🔹 Convert NaN / "nan" → None (Excel blank)
def clean_excel_value(value):
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, str) and value.strip().lower() == "nan":
        return None
    return value


@router.get("/Site-Status-download")
def download_site_Status(background_tasks: BackgroundTasks):

    if not os.path.exists(SS_TEMPLATE_PATH):
        raise HTTPException(status_code=500, detail="Template Excel not found")

    os.makedirs(TEMP_DIR, exist_ok=True)

    temp_file = os.path.join(TEMP_DIR, f"{uuid4().hex}.xlsx")

    # 1️⃣ Copy template
    shutil.copy(SS_TEMPLATE_PATH, temp_file)

    wb = load_workbook(temp_file)
    ws = wb.active

    # 2️⃣ Read Excel headers → DB columns
    excel_columns = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header in EXCEL_COLUMN_MAP:
            excel_columns[col] = EXCEL_COLUMN_MAP[header]

    # 3️⃣ Fetch DB data
    db = SessionLocal()
    try:
        rows = db.query(Data).all()
    finally:
        db.close()

    # 4️⃣ Write DB data → Excel (NULL safe)
    row_idx = 2
    for row in rows:
        for col_idx, db_field in excel_columns.items():
            raw_value = getattr(row, db_field, None)
            ws.cell(
                row=row_idx,
                column=col_idx
            ).value = clean_excel_value(raw_value)
        row_idx += 1
    wb.save(temp_file)
    background_tasks.add_task(delete_file, temp_file)
    return FileResponse(
        path=temp_file,
        filename=os.path.basename(SS_TEMPLATE_PATH),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
