from fastapi import APIRouter, HTTPException, BackgroundTasks
from fastapi.responses import FileResponse
from database import SessionLocal
from models import Data
from openpyxl import load_workbook
import shutil
import os
from uuid import uuid4

router = APIRouter()

# Excel Header -> DB field mapping (ONLY existing Excel headers)
EXCEL_COLUMN_MAP = {
    "Circle": "Circle_Name",
    "Project": "Project_Name",
    "Site ID": "Site_ID",
    "Vendor": "BTS_Vendor",
    "NSS ID": "NSS_ID",
    "MW OSM": "mw_osm",
    "Optics OSM": "optics_osm",
    "IP CR": "code_cr_details_ip_cr",

}

P_TEMPLATE_PATH = r"Provisioning\Sample_Provisoning_Update_Template (6).xlsx"
TEMP_DIR = r"C:\CubeXo Internship\FastAPI\CLP_API\temp_downloads"

def delete_file(path: str):
    if os.path.exists(path):
        os.remove(path)

@router.get("/Provisioning-Download", tags=["Proviosining-Download"])
def download_ip_provisioning(background_tasks: BackgroundTasks):

    if not os.path.exists(P_TEMPLATE_PATH):
        raise HTTPException(status_code=500, detail="Template Excel not found")

    os.makedirs(TEMP_DIR, exist_ok=True)

    temp_file = os.path.join(
        TEMP_DIR, f"{uuid4().hex}.xlsx"
    )

    # 1️⃣ Copy original Excel
    shutil.copy(P_TEMPLATE_PATH, temp_file)

    wb = load_workbook(temp_file)
    ws = wb.active

    # 2️⃣ Read Excel headers
    excel_columns = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header in EXCEL_COLUMN_MAP:
            excel_columns[col] = EXCEL_COLUMN_MAP[header]

    # 3️⃣ Fetch DB data
    db = SessionLocal()
    try:
        rows = db.query(Data).all()
    finally:
        db.close()

    # 4️⃣ Write data (ONLY mapped columns)
    row_idx = 2
    for row in rows:
        for col_idx, db_field in excel_columns.items():
            ws.cell(row=row_idx, column=col_idx).value = getattr(row, db_field, None)
        row_idx += 1

    wb.save(temp_file)

    # 5️⃣ Delete temp file AFTER download
    background_tasks.add_task(delete_file, temp_file)

    # 6️⃣ Download with SAME filename as template
    return FileResponse(
        path=temp_file,
        filename=os.path.basename(P_TEMPLATE_PATH),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
