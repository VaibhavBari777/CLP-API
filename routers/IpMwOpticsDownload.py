from fastapi.responses import FileResponse
from openpyxl import load_workbook
import os, shutil
from uuid import uuid4
from typing import List
from fastapi import APIRouter
from fastapi import HTTPException
from database import SessionLocal
from sqlalchemy import select
from models import Data

router = APIRouter(prefix="/download")
BASE_TEMPLATE_DIR = r"C:\CubeXo Internship\FastAPI\CLP_API\excel_outputs"
TEMP_DIR = r"C:\CubeXo Internship\FastAPI\CLP_API\temp_downloads"
os.makedirs(TEMP_DIR, exist_ok=True)

FILES = {
    "ip": "IP-sample-planning-circuit (26).xlsx",
    "mw": "MW-sample-planning-circuit (10).xlsx",
    "optics": "Optics-sample-planning-circuit (7).xlsx"
}

HEADER_ROW = 3
START_ROW = 4

# ---------- IP helper ----------
def create_temp_ip_excel(nss_ids: List[str]) -> str:
    original_path = os.path.join(BASE_TEMPLATE_DIR, FILES["ip"])
    temp_path = os.path.join(TEMP_DIR, f"ip_{uuid4().hex}.xlsx")
    shutil.copy(original_path, temp_path)
    wb = load_workbook(temp_path)
    ws = wb.active
    nss_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=HEADER_ROW, column=col).value and "NSS ID" in str(ws.cell(row=HEADER_ROW, column=col).value):
            nss_col = col
            break
    if not nss_col:
        raise Exception("NSS ID column not found in IP Excel")
    for row in range(START_ROW, ws.max_row + 1):
        ws.cell(row=row, column=nss_col).value = None
    for idx, nss in enumerate(nss_ids):
        ws.cell(row=START_ROW + idx, column=nss_col).value = nss
    wb.save(temp_path)
    return temp_path

# ---------- Generic helper ----------
def create_temp_generic_excel(file_type: str, nss_ids: List[str]) -> str:
    original_path = os.path.join(BASE_TEMPLATE_DIR, FILES[file_type])
    temp_path = os.path.join(TEMP_DIR, f"{file_type}_{uuid4().hex}.xlsx")
    shutil.copy(original_path, temp_path)
    wb = load_workbook(temp_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=START_ROW, max_col=1):
        row[0].value = None
    for idx, nss in enumerate(nss_ids, start=START_ROW):
        ws.cell(row=idx, column=1).value = nss
    wb.save(temp_path)
    return temp_path

@router.get("/ip-download", tags=["IP-Download"])
def ip_download():
    db = SessionLocal()
    try:
        nss_ids = db.execute(select(Data.NSS_ID).where(Data.NSS_ID.isnot(None))).scalars().all()
    finally:
        db.close()
    if not nss_ids:
        raise HTTPException(status_code=400, detail="No NSS_ID found")
    temp_file = create_temp_ip_excel(nss_ids)
    return FileResponse(path=temp_file, filename=FILES["ip"], media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ---------- MW route ----------
@router.get("/mw", tags=["MW-Download"])
def download_mw():
    db = SessionLocal()
    try:
        nss_ids = db.execute(select(Data.NSS_ID).where(Data.NSS_ID.isnot(None))).scalars().all()
    finally:
        db.close()
    if not nss_ids:
        raise HTTPException(status_code=400, detail="No NSS_ID found")
    temp_file = create_temp_generic_excel("mw", nss_ids)
    return FileResponse(path=temp_file, filename=FILES["mw"], media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ---------- Optics route ----------
@router.get("/optics", tags=["Optics-Download"])
def download_optics():
    db = SessionLocal()
    try:
        nss_ids = db.execute(select(Data.NSS_ID).where(Data.NSS_ID.isnot(None))).scalars().all()
    finally:
        db.close()
    if not nss_ids:
        raise HTTPException(status_code=400, detail="No NSS_ID found")
    temp_file = create_temp_generic_excel("optics", nss_ids)
    return FileResponse(path=temp_file, filename=FILES["optics"], media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


